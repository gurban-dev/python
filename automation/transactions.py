'''
Assume the following is in an excel spreadsheet:

1              A            B       C         D

2 transaction_id   product_id   price

3           1001            1   $5.95   =C2*0.9

4           1002            2   $6.95

5           1003            3   $7.95

Updating hundreds or thousands of excel spreadsheets can be
done much more quickly with Python automation.

Suppose the price needed to be reduced by ten percent.
Scrolling through all of the rows to apply a formula
to adjust the price would take a lot of time.
This process can be automated.

What if you wanted to open hundreds or thousands of excel
spreadsheets and add a chart to each one, this would take
a week or two. With Python automation, this process can be
completed in a matter of seconds.

Define a function, so that we can automate the
process of updating thousands of spreadsheets.
'''

'''
Install the openpyxl package with pip:
pip install openpyxl

An asterisk ( * ) to the left of a cell indicates that the
cell is still running and hasn't finished completion yet.
'''

# Instead of typing "openpyxl.", we can type "xl.".

# "xl" is an alias for the "openpyxl" package.
import openpyxl as xl

'''
From the "openpyxl" package, we have a module called
"chart", and from this module, two classes are being
imported: "BarChart" and "Reference".
'''
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    # The openpyxl.load_workbook() function takes in the filename
    # and returns a value of the workbook data type.
    
    # The variable "wb" is of the workbook data type.
    wb = xl.load_workbook(filename)
    
    '''
    If the transactions.xlsx excel spreadsheet were to be
    opened, it would become apparent that there is only
    "Sheet1" in the bottom left corner.
    
    "sheet1" is a Worksheet object.
    '''
    sheet1 = wb['Sheet1']
    
    # Accesses the coordinate 'A1' where "A" is
    # the column name and "1" is the row number.
    # cell = sheet1['a1']
    
    # print('\ncell:', cell)
    
    # sheet.cell(row, column) accesses row 1
    # and column 1 in the excel spreadsheet.
    # cell = sheet1.cell(1, 1)
    
    # print('\ncell.value:', cell.value)
    
    # Find the maximum amount of rows there are in the
    # excel spreadsheet.
    # print('sheet1.max_row:', sheet1.max_row)
    
    # This program must access rows 2, 3, and 4 in
    # the transactions.xlsx excel spreadsheet.
    
    # The range() function will generate values starting from 2
    # all the way up to, but not including sheet.max_row + 1.
    
    # Generate a range of values from 2 to 5, but not including 5.
    # for i in range(2, 5):
        # i = 2
        # i = 3
        # i = 4
    
    for row in range(2, sheet1.max_row + 1):
        # First iteration:
        # row = 2
        cell = sheet1.cell(row, 3)
                
        print(f'\nrow: {row}, cell.value: {cell.value}')
        
        # Reduce the price in the third column by ten percent.
        corrected_price = cell.value * 0.9
        
        # Get a reference to the fourth column in every row
        # by invoking cell(row, 4) on the "sheet" object.
        
        # First iteration:
        # row = 2
        corrected_price_cell = sheet1.cell(row, 4)
        
        # Add the corrected price to the newly added fourth
        # column by invoking the "value" property.
        
        # '$' mustn't be prepended or inserted before the value
        # of the "corrected_price_formatted" variable because then
        # the excel spreadsheet will think a string rather than a
        # floating-point number is being inserted.
        corrected_price_formatted = format(corrected_price, '.2f')
        
        # Cast the string returned by format() to a floating-point number.
        # The corrected price is being added to the column D in the excel
        # spreadsheet with the "value" property invoked.
        corrected_price_cell.value = float(corrected_price_formatted)

    '''
    Add a chart to the transaction.xlsx excel spreadsheet.
    
    The chart needs to obtain values.
    
    Use the "Reference" class to select a range of values
    from the excel sheet (rows 2-4 in this case).
    
    Create an instance of the "Reference" class
    and store a variable in it called "values".
    
    Select the cells in rows 2, 3, and 4 in column 4 or D.
    
    The "values" variable will be assigned the values from
    the fourth column.

		Documentation for openpyxl.chart.reference.Reference:

		https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.chart.reference.html
    '''
    values = Reference(sheet1,
			min_row=2,
			max_row=sheet1.max_row,
			min_col=4,
			max_col=4
    )
    
    print(f'\nvalues: {values}')
    
    print(f'\ntype(values): {type(values)}')
    
    '''
    Create an instance of the "BarChart" class
    and store it in a variable called "chart".
    
    BarChart is a class and in the subsequent line an instance
    or object of the BarChart class is being created.
    '''
    chart = BarChart()
    
    # Add the data from the fourth column to the "chart"
    # instance.
    chart.add_data(values)
    
    # Add the instance "chart" to the excel spreadsheet
    # in coordinate 'e2' or column E row 2.
    sheet1.add_chart(chart, 'e2')
    
    # Save the excel spreadsheet.
    wb.save(filename)

'''
If you are not sure whether your program has any bugs,
save the work in another file to avoid overwriting the
original excel spreadsheet.

The process_workbook() function has the competency to
process hundreds or thousands of spreadsheets.

With this program, the name of each file in a directory
can be obtained and passed to the process_workbook()
function. The process_workbook() function will update
each spreadsheet in just a second or less.

Be sure to use a .ipynb file rather than a .py for
running this program in Jupyter notebook.

If you use a .py file, 'transactions.xlsx' will not
be found.
'''

process_workbook('transactions.xlsx')