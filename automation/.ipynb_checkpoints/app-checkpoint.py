# Instead of typing "openpyxl.", we can type "xl.".
# "xl" is an alias for the "openpyxl" package.
import openpyxl as xl

'''
From the "openpyxl" package, we have a module called
"chart", and from this module, two classes are being
imported: "BarChart" and "Reference".
'''
from openpyxl.chart import BarChart, Reference

'''
3:57:50
https://www.youtube.com/watch?v=_uQrJ0TkZlc

Assume the following is in an excel spreadsheet:
1              A            B       C         D
2 transaction_id   product_id   price
3           1001            1   $5.95   =C2*0.9
4           1002            2   $6.95
5           1003            3   $7.95

Updating hundreds or thousands of excel spreadsheets can be
done much more quickly with Python automation.

Suppose the price needed to be reduced by ten percent.
Scrolling through all of the rows to apply a formula
to adjust the price would take a lot of time.
This process can be automated.

What if you wanted to open hundreds or thousands of excel
spreadsheets and add a chart to each one, this would take
a week or two. With Python automation, this process can be
completed in a matter of seconds.
'''

# Define a function, so that we can automate the
# process of updating thousands of spreadsheets.
def process_workbook(filename):
  # Loads an excel work book and returns a work book object.
  wb = xl.load_workbook(filename)
	sheet = wb['Sheet1'] # Accesses the first sheet of the work book.
	# cell = sheet['a1'] # Accesses the coordinate 'a1'.
	# cell = sheet.cell(1, 1) # sheet.cell(row, column) Accesses row 1 column 1.

	# print('cell.value:', cell.value)

	# print('sheet.max_row:', sheet.max_row) # Find the maximum amount of rows there are in the spreadsheet.

	for row in range(2, sheet.max_row + 1):
			cell = sheet.cell(row, 3)
			# print('row:', row)
			# print('cell.value:', cell.value)
			corrected_price = cell.value * 0.9

			# Add a new column with the "corrected_price".
			corrected_price_cell = sheet.cell(row, 4)
			corrected_price_cell.value = corrected_price

	# Add a chart to the current sheet.
	# First, use the "Reference" class to select a range of values from the excel sheet (rows 2-4).
	# We are creating an instance of the "Reference" class and are storing it in a variable called values.
	values = Reference(sheet, 
					min_row=2, 
					max_row=sheet.max_row,
					min_col=4, # On lines 33 & 34 we are limiting the range of
					max_col=4  # the cells we are selecting to the fourth column. 
					)
	# The "values" variable will have all of the values in the fourth column.

	chart = BarChart() # Create an instance of the "BarChart" class and store it in a variable called "chart".
	chart.add_data(values) # Add the data in the fourth column to the chart instance.
	sheet.add_chart(chart, 'e2') # Add the chart instance to the excel sheet in coordinate 'e2'.

	wb.save(filename)

# If you are not sure whether your program has any bugs, save the work in another file to avoid
# overwriting the original excel sheet.

# We can reuse the 'process_workbook' function to process thousands of spreadsheets.
# With this program we can get each file in a directory, and pass the name of the file
# to the 'process_workbook' function. The 'process_workbook' function will update each
# spreadsheet in just a second or less.
